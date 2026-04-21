from __future__ import annotations

import io
import re
import json
import html
import unicodedata
import calendar
from collections import defaultdict
from datetime import date

import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


REQUIRED_COLS = ["Estado"]
OPTIONAL_COLS = ["Nombre", "Marc.", "NvoEstado"]

MONTH_NAMES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}


# =========================
# UI
# =========================
def inject_css() -> None:
    st.markdown(
        """
        <style>
        :root{
            --unsj-blue-1:#002B63;
            --unsj-blue-2:#0A58CA;
            --unsj-blue-3:#79B8FF;
            --glass-1: rgba(255,255,255,.10);
            --glass-2: rgba(255,255,255,.05);
            --glass-br: rgba(255,255,255,.14);
            --txt: #F7FBFF;
            --txt-soft: rgba(247,251,255,.78);
            --shadow: 0 14px 40px rgba(0,0,0,.22);
        }

        html, body, [class*="css"] {
            color: var(--txt);
        }

        .stApp {
            background:
                radial-gradient(circle at 14% 18%, rgba(121,184,255,.24), transparent 24%),
                radial-gradient(circle at 85% 15%, rgba(10,88,202,.22), transparent 20%),
                radial-gradient(circle at 78% 80%, rgba(0,43,99,.24), transparent 25%),
                linear-gradient(135deg, #02101f 0%, #062349 38%, #0B3971 70%, #124E93 100%);
        }

        .block-container {
            max-width: 1180px;
            padding-top: 1rem;
            padding-bottom: 1.4rem;
        }

        header, footer {visibility: hidden;}
        div[data-testid="stToolbar"] {visibility: hidden; height: 0px;}

        .hero-wrap{
            background: linear-gradient(180deg, rgba(255,255,255,.10), rgba(255,255,255,.05));
            border: 1px solid rgba(255,255,255,.14);
            border-radius: 28px;
            padding: 22px 26px;
            box-shadow: var(--shadow);
            backdrop-filter: blur(18px) saturate(160%);
            -webkit-backdrop-filter: blur(18px) saturate(160%);
            text-align:center;
            margin-bottom: 14px;
        }

        .hero-title{
            font-size: 2.15rem;
            font-weight: 900;
            letter-spacing: .8px;
            color: #FFFFFF;
            line-height: 1.02;
            text-transform: uppercase;
        }

        .hero-sub{
            margin-top: 8px;
            font-size: .88rem;
            color: rgba(255,255,255,.76);
            letter-spacing: .7px;
            text-transform: uppercase;
        }

        .toolbar-wrap{
            background: linear-gradient(180deg, rgba(255,255,255,.08), rgba(255,255,255,.04));
            border: 1px solid rgba(255,255,255,.12);
            border-radius: 24px;
            padding: 14px 16px;
            box-shadow: 0 10px 30px rgba(0,0,0,.18);
            backdrop-filter: blur(16px) saturate(160%);
            -webkit-backdrop-filter: blur(16px) saturate(160%);
            margin-bottom: 14px;
        }

        .kpi {
            border: 1px solid rgba(255,255,255,0.14);
            border-radius: 22px;
            padding: 16px 16px;
            min-height: 110px;
            background: linear-gradient(180deg, rgba(255,255,255,.10), rgba(255,255,255,.05));
            box-shadow: 0 10px 28px rgba(0,0,0,.16);
            backdrop-filter: blur(14px) saturate(155%);
            -webkit-backdrop-filter: blur(14px) saturate(155%);
        }

        .kpi .label {
            opacity: .82;
            font-size: .86rem;
            font-weight: 700;
            color: rgba(255,255,255,.80);
            text-transform: uppercase;
            letter-spacing: .5px;
        }

        .kpi .value {
            font-size: 1.72rem;
            font-weight: 900;
            line-height: 1.06;
            margin-top: 8px;
            color:#FFFFFF;
            text-transform: uppercase;
        }

        .kpi .sub {
            opacity: .72;
            font-size: .82rem;
            margin-top: .35rem;
            color: rgba(255,255,255,.72);
            text-transform: uppercase;
            letter-spacing: .35px;
        }

        .pill {
            display:inline-block;
            padding:8px 12px;
            border-radius:999px;
            border:1px solid rgba(255,255,255,.14);
            background:linear-gradient(180deg, rgba(255,255,255,.10), rgba(255,255,255,.05));
            font-size:.80rem;
            color: rgba(255,255,255,.92);
            box-shadow: 0 8px 22px rgba(0,0,0,.14);
            backdrop-filter: blur(14px);
            -webkit-backdrop-filter: blur(14px);
            text-transform: uppercase;
            letter-spacing: .35px;
        }

        .hr {
            height:1px;
            background: linear-gradient(90deg, transparent, rgba(255,255,255,.22), transparent);
            margin: 1rem 0 1rem 0;
        }

        .calendar-wrap{
            background: linear-gradient(180deg, rgba(255,255,255,.07), rgba(255,255,255,.04));
            border: 1px solid rgba(255,255,255,.12);
            border-radius: 20px;
            padding: 14px;
            margin-top: 12px;
        }

        .calendar-weekday{
            text-align:center;
            font-size:.78rem;
            font-weight:800;
            color: rgba(255,255,255,.78);
            letter-spacing:.45px;
            margin-bottom:6px;
        }

        .calendar-note{
            font-size:.80rem;
            opacity:.82;
            letter-spacing:.25px;
            text-transform:uppercase;
        }

        div[data-testid="stFileUploader"] > section,
        div[data-testid="stDataFrame"],
        div[data-testid="stTable"],
        div[data-testid="stDataEditor"],
        div[data-testid="stMetric"],
        div[data-testid="stVerticalBlockBorderWrapper"],
        div[data-testid="stAlert"],
        div[data-testid="stExpander"] {
            border-radius: 22px !important;
        }

        div[data-testid="stFileUploader"] > section {
            background: linear-gradient(180deg, rgba(255,255,255,.10), rgba(255,255,255,.05)) !important;
            border: 1px solid rgba(255,255,255,.14) !important;
            box-shadow: 0 10px 28px rgba(0,0,0,.15) !important;
            backdrop-filter: blur(16px) saturate(160%);
            -webkit-backdrop-filter: blur(16px) saturate(160%);
        }

        .stButton > button,
        .stDownloadButton > button {
            width: 100%;
            border-radius: 16px !important;
            border: 1px solid rgba(255,255,255,.16) !important;
            background: linear-gradient(180deg, rgba(121,184,255,.24), rgba(10,88,202,.16)) !important;
            color: white !important;
            font-weight: 900 !important;
            min-height: 46px !important;
            box-shadow: 0 10px 24px rgba(0,0,0,.18) !important;
            backdrop-filter: blur(14px);
            -webkit-backdrop-filter: blur(14px);
            text-transform: uppercase !important;
            letter-spacing: .4px !important;
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover {
            border-color: rgba(255,255,255,.24) !important;
            transform: translateY(-1px);
            box-shadow: 0 14px 28px rgba(0,0,0,.22) !important;
        }

        div[data-baseweb="select"] > div,
        .stTextInput > div > div > input,
        .stTextArea textarea,
        input[type="number"] {
            background: rgba(255,255,255,.08) !important;
            border: 1px solid rgba(255,255,255,.15) !important;
            border-radius: 14px !important;
            color: white !important;
            backdrop-filter: blur(12px);
            -webkit-backdrop-filter: blur(12px);
            text-transform: uppercase;
        }

        div[data-baseweb="tab-list"] {
            gap: 10px;
            background: transparent !important;
        }

        button[data-baseweb="tab"] {
            border-radius: 16px !important;
            padding: 10px 18px !important;
            background: rgba(255,255,255,.06) !important;
            border: 1px solid rgba(255,255,255,.12) !important;
            color: rgba(255,255,255,.88) !important;
            backdrop-filter: blur(14px);
            -webkit-backdrop-filter: blur(14px);
            text-transform: uppercase;
            letter-spacing: .35px;
        }

        button[data-baseweb="tab"][aria-selected="true"] {
            background: linear-gradient(180deg, rgba(121,184,255,.24), rgba(10,88,202,.15)) !important;
            border-color: rgba(255,255,255,.22) !important;
            color: #fff !important;
            box-shadow: 0 10px 24px rgba(0,0,0,.16);
        }

        [data-testid="stDataFrame"] > div,
        [data-testid="stDataEditor"] > div {
            background: linear-gradient(180deg, rgba(255,255,255,.10), rgba(255,255,255,.05)) !important;
            border: 1px solid rgba(255,255,255,.14) !important;
            box-shadow: 0 10px 28px rgba(0,0,0,.15) !important;
            backdrop-filter: blur(14px);
            -webkit-backdrop-filter: blur(14px);
        }

        [data-testid="stMarkdownContainer"] p,
        [data-testid="stMarkdownContainer"] li,
        .stCaption {
            color: rgba(255,255,255,.88);
        }

        label, .st-emotion-cache-16txtl3, .st-emotion-cache-pkbazv {
            text-transform: uppercase !important;
            letter-spacing: .35px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def hero_header() -> None:
    st.markdown(
        """
        <div class="hero-wrap">
            <div class="hero-title">CONTROL DE ASISTENCIA APP</div>
            <div class="hero-sub">DESARROLLADA POR MARCELO JAMESON</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def kpi_card(label: str, value: str, sub: str = "") -> None:
    st.markdown(
        f"""
        <div class="kpi">
            <div class="label">{html.escape(str(label))}</div>
            <div class="value">{html.escape(str(value))}</div>
            <div class="sub">{html.escape(str(sub))}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


# =========================
# COPY
# =========================
def copy_table_button(df: pd.DataFrame, label: str, key: str) -> None:
    if df is None:
        df = pd.DataFrame()

    tsv = df.to_csv(sep="\t", index=False)
    payload = {"tsv": tsv}
    j = json.dumps(payload)

    st.components.v1.html(
        f"""
        <div style="display:flex; gap:10px; align-items:center; margin: 8px 0 10px 0;">
          <button id="btn_{key}" style="
              padding:12px 14px; border-radius:16px; border:1px solid rgba(255,255,255,.16);
              background:linear-gradient(180deg, rgba(121,184,255,.24), rgba(10,88,202,.16));
              color:white; font-weight:900; cursor:pointer; width:100%;
              box-shadow:0 10px 24px rgba(0,0,0,.18);
              backdrop-filter: blur(14px);
              text-transform:uppercase; letter-spacing:.35px;
          ">{html.escape(label)}</button>
          <span id="ok_{key}" style="opacity:.0; font-weight:800; color:white;">COPIADO ✅</span>
        </div>

        <script>
        const data_{key} = {j};
        const btn_{key} = document.getElementById("btn_{key}");
        const ok_{key} = document.getElementById("ok_{key}");

        btn_{key}.addEventListener("click", async () => {{
          try {{
            await navigator.clipboard.writeText(data_{key}.tsv);
            ok_{key}.style.opacity = "1";
            setTimeout(() => ok_{key}.style.opacity = "0", 1400);
          }} catch (e) {{
            alert("No se pudo copiar automáticamente. Probá con Chrome/Edge o habilitá permisos de portapapeles.");
          }}
        }});
        </script>
        """,
        height=66,
    )


# =========================
# HELPERS
# =========================
def minutes_to_hhmm(mins: int) -> str:
    mins = int(round(mins)) if mins is not None else 0
    h = mins // 60
    m = mins % 60
    return f"{h:02d}:{m:02d}"


def delta_short(mins: int) -> str:
    mins = int(round(mins))
    sign = "+" if mins > 0 else "-" if mins < 0 else ""
    mins_abs = abs(mins)
    h, m = mins_abs // 60, mins_abs % 60
    if h == 0:
        return f"{sign}{m:02d}M" if sign else "0M"
    return f"{sign}{h}H {m:02d}M" if sign else f"{h}H {m:02d}M"


def normalize_text_key(value: str) -> str:
    value = str(value or "").strip().upper()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = re.sub(r"\s+", " ", value)
    return value


def display_dni(value: str) -> str:
    v = str(value or "").strip()
    return v if v else "SIN DNI"


def read_excel_auto(file) -> pd.DataFrame:
    try:
        return pd.read_excel(file)
    except Exception:
        file.seek(0)
        try:
            return pd.read_excel(file, engine="openpyxl")
        except Exception:
            file.seek(0)
            return pd.read_excel(file, engine="xlrd")


def validate_format(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    missing_required = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing_required:
        raise ValueError(
            f"FORMATO INCORRECTO DEL RELOJ. FALTA LA COLUMNA OBLIGATORIA: {', '.join(missing_required)}"
        )

    for col in OPTIONAL_COLS:
        if col not in df.columns:
            df[col] = ""

    return df


def parse_and_clean(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    for col in ["Nombre", "Marc.", "Estado", "NvoEstado"]:
        if col not in df.columns:
            df[col] = ""

    df["__rowid__"] = range(1, len(df) + 1)

    df["DNI"] = (
        df["Nombre"]
        .fillna("")
        .astype(str)
        .str.replace(r"\D", "", regex=True)
        .str.strip()
    )

    df["Empleado"] = (
        df["Marc."]
        .fillna("")
        .astype(str)
        .replace({"nan": "", "None": ""})
        .str.strip()
    )

    df["FechaHora"] = pd.to_datetime(df["Estado"], errors="coerce", dayfirst=True)
    df = df.dropna(subset=["FechaHora"]).copy()

    def resolve_employee_name(row) -> str:
        emp = str(row["Empleado"] or "").strip()
        dni = str(row["DNI"] or "").strip()
        rowid = int(row["__rowid__"])

        if emp:
            return emp
        if dni:
            return f"SIN NOMBRE · DNI {dni}"
        return f"SIN IDENTIFICAR · REG {rowid}"

    df["Empleado"] = df.apply(resolve_employee_name, axis=1)

    def resolve_employee_key(row) -> str:
        dni = str(row["DNI"] or "").strip()
        emp = str(row["Empleado"] or "").strip()
        rowid = int(row["__rowid__"])

        if dni:
            return f"DNI::{dni}"
        if emp and not emp.startswith("SIN IDENTIFICAR · REG "):
            return f"NOMBRE::{normalize_text_key(emp)}"
        return f"REG::{rowid}"

    df["EmployeeKey"] = df.apply(resolve_employee_key, axis=1)
    df["Fecha"] = df["FechaHora"].dt.date

    df = df.sort_values(["Empleado", "DNI", "FechaHora"]).reset_index(drop=True)
    df["DNI"] = df["DNI"].astype(str)
    return df


def init_profiles(raw: pd.DataFrame) -> pd.DataFrame:
    base = (
        raw[["EmployeeKey", "DNI", "Empleado"]]
        .drop_duplicates()
        .sort_values(["Empleado", "DNI"])
        .reset_index(drop=True)
    )

    if "profiles" not in st.session_state:
        p = base.copy()
        p["Tipo"] = "NO Docente"
        st.session_state["profiles"] = p
        return p

    p = st.session_state["profiles"].copy()
    merged = base.merge(p[["EmployeeKey", "Tipo"]], on="EmployeeKey", how="left")
    merged["Tipo"] = merged["Tipo"].fillna("NO Docente")
    merged = merged[["EmployeeKey", "DNI", "Empleado", "Tipo"]]
    st.session_state["profiles"] = merged
    return merged


def apply_profiles(raw: pd.DataFrame, profiles: pd.DataFrame) -> pd.DataFrame:
    m = raw.merge(profiles[["EmployeeKey", "Tipo"]], on="EmployeeKey", how="left")
    m["Tipo"] = m["Tipo"].fillna("NO Docente")
    return m


def parse_holidays(text: str) -> set[date]:
    holidays: set[date] = set()
    if not text:
        return holidays

    for line in str(text).replace(",", "\n").splitlines():
        raw = line.strip()
        if not raw:
            continue
        try:
            holidays.add(pd.to_datetime(raw, dayfirst=True).date())
        except Exception:
            pass
    return holidays


def holidays_to_text(holidays: set[date]) -> str:
    if not holidays:
        return ""
    return "\n".join(sorted([d.strftime("%d/%m/%Y") for d in holidays]))


def init_holidays_state() -> None:
    if "holidays_set" not in st.session_state:
        st.session_state["holidays_set"] = set()
    if "holidays_text" not in st.session_state:
        st.session_state["holidays_text"] = ""
    if "holiday_calendar_month" not in st.session_state:
        today = date.today()
        st.session_state["holiday_calendar_month"] = today.month
    if "holiday_calendar_year" not in st.session_state:
        st.session_state["holiday_calendar_year"] = date.today().year


def sync_holidays_text_from_set() -> None:
    st.session_state["holidays_text"] = holidays_to_text(st.session_state["holidays_set"])


def apply_text_holidays() -> None:
    st.session_state["holidays_set"] = parse_holidays(st.session_state.get("holidays_text", ""))
    sync_holidays_text_from_set()


def clear_all_holidays() -> None:
    st.session_state["holidays_set"] = set()
    sync_holidays_text_from_set()


def toggle_holiday(day_value: date) -> None:
    holidays = set(st.session_state["holidays_set"])
    if day_value in holidays:
        holidays.remove(day_value)
    else:
        holidays.add(day_value)
    st.session_state["holidays_set"] = holidays
    sync_holidays_text_from_set()


def render_holiday_calendar() -> None:
    month = int(st.session_state["holiday_calendar_month"])
    year = int(st.session_state["holiday_calendar_year"])
    holidays = st.session_state["holidays_set"]

    st.markdown('<div class="calendar-wrap">', unsafe_allow_html=True)

    nav1, nav2, nav3, nav4 = st.columns([1.2, 1, 1, 1.4])
    with nav1:
        selected_month_name = st.selectbox(
            "MES",
            options=list(MONTH_NAMES.values()),
            index=month - 1,
            key="holiday_month_name_selector",
            label_visibility="collapsed",
        )
        month = {v: k for k, v in MONTH_NAMES.items()}[selected_month_name]
        st.session_state["holiday_calendar_month"] = month

    with nav2:
        selected_year = st.number_input(
            "AÑO",
            min_value=2000,
            max_value=2100,
            value=year,
            step=1,
            key="holiday_year_number",
            label_visibility="collapsed",
        )
        year = int(selected_year)
        st.session_state["holiday_calendar_year"] = year

    with nav3:
        if st.button("HOY", key="holiday_go_today"):
            today = date.today()
            st.session_state["holiday_calendar_month"] = today.month
            st.session_state["holiday_calendar_year"] = today.year
            st.rerun()

    with nav4:
        st.markdown(
            f"""<div class="pill">{MONTH_NAMES[month]} {year}</div>""",
            unsafe_allow_html=True,
        )

    week_cols = st.columns(7)
    weekday_names = ["LUN", "MAR", "MIÉ", "JUE", "VIE", "SÁB", "DOM"]
    for col, wd in zip(week_cols, weekday_names):
        with col:
            st.markdown(f'<div class="calendar-weekday">{wd}</div>', unsafe_allow_html=True)

    cal = calendar.Calendar(firstweekday=0)
    weeks = cal.monthdayscalendar(year, month)

    for week_idx, week in enumerate(weeks):
        cols = st.columns(7)
        for day_idx, day_num in enumerate(week):
            with cols[day_idx]:
                if day_num == 0:
                    st.markdown("&nbsp;", unsafe_allow_html=True)
                    continue

                current_date = date(year, month, day_num)
                is_holiday = current_date in holidays

                label = f"🟦 {day_num}" if is_holiday else f"{day_num}"
                key = f"holiday_btn_{year}_{month}_{day_num}_{week_idx}_{day_idx}"

                if st.button(label, key=key, use_container_width=True):
                    toggle_holiday(current_date)
                    st.rerun()

    st.markdown(
        '<div class="calendar-note">TOCÁ UN DÍA PARA AGREGARLO O QUITARLO COMO FERIADO.</div>',
        unsafe_allow_html=True,
    )
    st.markdown('</div>', unsafe_allow_html=True)


def is_special_day(day_value, holidays: set[date]) -> bool:
    ts = pd.to_datetime(day_value)
    return ts.weekday() >= 5 or ts.date() in holidays


def pair_alternating(times: list[pd.Timestamp]) -> tuple[int, int]:
    times = [t for t in times if pd.notna(t)]
    times.sort()
    total = 0
    pairs = 0
    for i in range(0, len(times) - 1, 2):
        a, b = times[i], times[i + 1]
        if b >= a:
            total += int((b - a).total_seconds() // 60)
            pairs += 1
    return total, pairs


def split_interval_by_day(start: pd.Timestamp, end: pd.Timestamp) -> list[tuple[pd.Timestamp, int]]:
    chunks = []
    current = start

    while current < end:
        next_midnight = (current.normalize() + pd.Timedelta(days=1))
        chunk_end = min(next_midnight, end)
        minutes = int((chunk_end - current).total_seconds() // 60)
        if minutes > 0:
            chunks.append((current.normalize(), minutes))
        current = chunk_end

    return chunks


# =========================
# CÁLCULO NORMAL
# =========================
def calc_daily_standard(raw: pd.DataFrame, expected_nodoc: int, holidays: set[date]) -> pd.DataFrame:
    rows = []

    for (ekey, dni, emp, tipo, day), g in raw.groupby(
        ["EmployeeKey", "DNI", "Empleado", "Tipo", "Fecha"], dropna=False
    ):
        g = g.sort_values("FechaHora")
        times = g["FechaHora"].tolist()
        marc = int(g.shape[0])

        first = times[0] if times else pd.NaT
        last = times[-1] if times else pd.NaT

        worked_pairs, pairs = pair_alternating(times)

        span = 0
        if pd.notna(first) and pd.notna(last) and last >= first:
            span = int((last - first).total_seconds() // 60)

        fecha_ts = pd.to_datetime(day)
        day_is_special = is_special_day(fecha_ts.date(), holidays)
        day_type = "FERIADO/FIN DE SEMANA" if day_is_special else "HÁBIL"

        incompleto = (marc < 2) if tipo == "NO Docente" else (pairs == 0)
        cortes = (pairs >= 2)

        if tipo == "Docente":
            worked = worked_pairs
            expected = 0
            saldo = 0
            cumple = ""
        else:
            worked = span if (marc >= 2 and pd.notna(first) and pd.notna(last)) else 0

            if day_is_special:
                expected = 0
                saldo = worked
                if worked > 0 and not incompleto:
                    cumple = "EXTRA"
                elif worked > 0 and incompleto:
                    cumple = "INCOMPLETO"
                else:
                    cumple = ""
            else:
                expected = expected_nodoc if marc >= 1 else 0
                saldo = worked - expected if expected else 0

                if expected and not incompleto:
                    cumple = "OK" if saldo >= 0 else "FALTA"
                elif expected and incompleto:
                    cumple = "INCOMPLETO"
                else:
                    cumple = ""

        rows.append(
            {
                "EmployeeKey": ekey,
                "DNI": str(dni),
                "Empleado": emp,
                "Tipo": tipo,
                "Fecha": fecha_ts,
                "Tipo_dia": day_type,
                "Es_fin_de_semana": "SI" if day_is_special else "",
                "Primera": first,
                "Ultima": last,
                "Horas": minutes_to_hhmm(worked),
                "Minutos": int(worked),
                "Esperado_min": int(expected),
                "Esperado": minutes_to_hhmm(expected),
                "Saldo_min": int(saldo),
                "Saldo": delta_short(saldo),
                "Cumple": cumple,
                "Marcaciones": marc,
                "Pares_estimados": int(pairs),
                "Cortes": "SI" if cortes else "",
                "Incompleto": "SI" if incompleto else "",
            }
        )

    d = pd.DataFrame(rows)
    if d.empty:
        return d
    return d.sort_values(["Tipo", "Empleado", "DNI", "Fecha"]).reset_index(drop=True)


# =========================
# CÁLCULO CHOFERES
# =========================
def calc_daily_drivers(raw: pd.DataFrame, expected_nodoc: int, holidays: set[date]) -> pd.DataFrame:
    rows = []

    for (ekey, dni, emp, tipo), g_emp in raw.groupby(
        ["EmployeeKey", "DNI", "Empleado", "Tipo"], dropna=False
    ):
        g_emp = g_emp.sort_values("FechaHora").copy()

        if tipo == "Docente":
            g_doc = calc_daily_standard(g_emp.copy(), expected_nodoc, holidays)
            if not g_doc.empty:
                rows.extend(g_doc.to_dict("records"))
            continue

        times = g_emp["FechaHora"].tolist()

        raw_day_stats = {}
        for day, g_day in g_emp.groupby("Fecha"):
            g_day = g_day.sort_values("FechaHora")
            raw_day_stats[day] = {
                "Marcaciones": int(g_day.shape[0]),
                "Primera": g_day["FechaHora"].iloc[0] if not g_day.empty else pd.NaT,
                "Ultima": g_day["FechaHora"].iloc[-1] if not g_day.empty else pd.NaT,
            }

        day_worked = defaultdict(int)
        day_expected = defaultdict(int)
        day_pairs = defaultdict(int)

        for i in range(0, len(times) - 1, 2):
            start = times[i]
            end = times[i + 1]
            if pd.isna(start) or pd.isna(end) or end < start:
                continue

            remaining_normal = expected_nodoc
            chunks = split_interval_by_day(start, end)

            for day_ts, minutes in chunks:
                day_date = day_ts.date()
                special = is_special_day(day_date, holidays)

                if special:
                    normal_chunk = 0
                else:
                    normal_chunk = min(minutes, remaining_normal)
                    remaining_normal -= normal_chunk

                day_worked[day_date] += minutes
                day_expected[day_date] += normal_chunk
                day_pairs[day_date] += 1

        unmatched_day = None
        if len(times) % 2 != 0:
            unmatched_day = pd.to_datetime(times[-1]).date()

        all_days = set(raw_day_stats.keys()) | set(day_worked.keys())

        for day in sorted(all_days):
            fecha_ts = pd.to_datetime(day)
            special = is_special_day(day, holidays)
            day_type = "FERIADO/FIN DE SEMANA" if special else "HÁBIL"

            marc = raw_day_stats.get(day, {}).get("Marcaciones", 0)
            first = raw_day_stats.get(day, {}).get("Primera", pd.NaT)
            last = raw_day_stats.get(day, {}).get("Ultima", pd.NaT)

            worked = int(day_worked.get(day, 0))
            expected = int(day_expected.get(day, 0))
            saldo = worked - expected
            pairs = int(day_pairs.get(day, 0))

            incompleto = unmatched_day == day
            cortes = marc >= 3

            if incompleto:
                cumple = "INCOMPLETO"
            else:
                if worked > 0 and expected == 0 and saldo > 0:
                    cumple = "EXTRA"
                elif expected > 0:
                    cumple = "OK" if saldo >= 0 else "FALTA"
                else:
                    cumple = ""

            rows.append(
                {
                    "EmployeeKey": ekey,
                    "DNI": str(dni),
                    "Empleado": emp,
                    "Tipo": tipo,
                    "Fecha": fecha_ts,
                    "Tipo_dia": day_type,
                    "Es_fin_de_semana": "SI" if special else "",
                    "Primera": first,
                    "Ultima": last,
                    "Horas": minutes_to_hhmm(worked),
                    "Minutos": int(worked),
                    "Esperado_min": int(expected),
                    "Esperado": minutes_to_hhmm(expected),
                    "Saldo_min": int(saldo),
                    "Saldo": delta_short(saldo),
                    "Cumple": cumple,
                    "Marcaciones": int(marc),
                    "Pares_estimados": int(pairs),
                    "Cortes": "SI" if cortes else "",
                    "Incompleto": "SI" if incompleto else "",
                }
            )

    d = pd.DataFrame(rows)
    if d.empty:
        return d
    return d.sort_values(["Tipo", "Empleado", "DNI", "Fecha"]).reset_index(drop=True)


def calc_daily(raw: pd.DataFrame, expected_nodoc: int, holidays: set[date], driver_mode: bool) -> pd.DataFrame:
    if driver_mode:
        return calc_daily_drivers(raw, expected_nodoc, holidays)
    return calc_daily_standard(raw, expected_nodoc, holidays)


# =========================
# CORRECCIÓN AUTOMÁTICA NO DOCENTE
# =========================
def correct_missing_punches_for_employee(raw_emp: pd.DataFrame, expected_nodoc: int) -> tuple[pd.DataFrame, int]:
    if raw_emp.empty:
        return raw_emp, 0

    corrected = raw_emp.copy()
    corrected["Fecha"] = corrected["FechaHora"].dt.date

    fixes = []
    nfix = 0
    for day, g in corrected.groupby("Fecha"):
        times = sorted(g["FechaHora"].tolist())
        if len(times) == 1:
            t = times[0]
            fix_out = t + pd.to_timedelta(expected_nodoc, unit="m")
            fixes.append({"FechaHora": fix_out, "Fecha": day})
            nfix += 1

    if fixes:
        fx_rows = []
        template = corrected.iloc[0].copy()
        for f in fixes:
            row = template.copy()
            row["FechaHora"] = f["FechaHora"]
            row["Fecha"] = f["Fecha"]
            row["NvoEstado"] = "AUTO_FIX"
            row["Estado"] = row["FechaHora"].strftime("%d/%m/%Y %H:%M")
            fx_rows.append(row)

        add = pd.DataFrame(fx_rows)
        corrected = (
            pd.concat([corrected, add], ignore_index=True)
            .sort_values("FechaHora")
            .reset_index(drop=True)
        )

    corrected["Fecha"] = corrected["FechaHora"].dt.date
    return corrected, nfix


def correct_missing_punches_all(raw: pd.DataFrame, expected_nodoc: int) -> tuple[pd.DataFrame, int]:
    if raw.empty:
        return raw, 0

    docentes = raw[raw["Tipo"] == "Docente"].copy()
    nodoc = raw[raw["Tipo"] == "NO Docente"].copy()
    if nodoc.empty:
        return raw, 0

    fixed_parts = []
    total_fixes = 0

    for ekey, g in nodoc.groupby(["EmployeeKey"]):
        corrected, nfix = correct_missing_punches_for_employee(g.copy(), expected_nodoc)
        fixed_parts.append(corrected)
        total_fixes += nfix

    nodoc_fixed = pd.concat(fixed_parts, ignore_index=True) if fixed_parts else nodoc
    out = pd.concat([docentes, nodoc_fixed], ignore_index=True)
    out = out.sort_values(["Empleado", "DNI", "FechaHora"]).reset_index(drop=True)
    return out, total_fixes


# =========================
# RESÚMENES
# =========================
def summarize(daily: pd.DataFrame) -> pd.DataFrame:
    if daily.empty:
        return pd.DataFrame()

    def extras_pos(x: pd.Series) -> int:
        return int(x[x > 0].sum())

    def faltas_pos(x: pd.Series) -> int:
        return int((-x[x < 0]).sum())

    s = (
        daily.groupby(["EmployeeKey", "Empleado", "DNI", "Tipo"], as_index=False)
        .agg(
            Dias=("Fecha", "nunique"),
            Total_min=("Minutos", "sum"),
            Prom_min=("Minutos", "mean"),
            Incompletos=("Incompleto", lambda x: int((x == "SI").sum())),
            Cortes=("Cortes", lambda x: int((x == "SI").sum())),
            Marcaciones=("Marcaciones", "sum"),
            Esperado_min=("Esperado_min", "sum"),
            Saldo_min=("Saldo_min", "sum"),
            Extras_min=("Saldo_min", extras_pos),
            Faltas_min=("Saldo_min", faltas_pos),
            Dias_OK=("Cumple", lambda x: int((x == "OK").sum())),
            Dias_FALTA=("Cumple", lambda x: int((x == "FALTA").sum())),
            Dias_INCOMPL=("Cumple", lambda x: int((x == "INCOMPLETO").sum())),
            Dias_EXTRA=("Cumple", lambda x: int((x == "EXTRA").sum())),
        )
        .sort_values(["Tipo", "Empleado", "DNI"])
        .reset_index(drop=True)
    )

    s["DNI"] = s["DNI"].astype(str)
    s["Total"] = s["Total_min"].round().astype(int).apply(minutes_to_hhmm)
    s["Prom/día"] = s["Prom_min"].round().astype(int).apply(minutes_to_hhmm)
    s["Extras"] = s["Extras_min"].apply(minutes_to_hhmm)
    s["Faltas"] = s["Faltas_min"].apply(minutes_to_hhmm)
    s["Saldo"] = s["Saldo_min"].apply(delta_short)

    def pct_row(r):
        if r["Tipo"] == "NO Docente" and r["Esperado_min"] > 0:
            return f"{(r['Total_min'] / r['Esperado_min'] * 100):.0f}%"
        return ""

    s["Cumplimiento"] = s.apply(pct_row, axis=1)

    cols = [
        "EmployeeKey",
        "Empleado", "DNI", "Tipo",
        "Dias",
        "Total", "Total_min",
        "Prom/día",
        "Esperado_min",
        "Extras", "Extras_min",
        "Faltas", "Faltas_min",
        "Saldo", "Saldo_min",
        "Cumplimiento",
        "Marcaciones",
        "Incompletos",
        "Cortes",
        "Dias_OK", "Dias_FALTA", "Dias_INCOMPL", "Dias_EXTRA",
    ]
    cols = [c for c in cols if c in s.columns]
    return s[cols]


def employee_detail_table(daily_emp: pd.DataFrame) -> pd.DataFrame:
    d = daily_emp.copy()
    d["Fecha"] = pd.to_datetime(d["Fecha"]).dt.date
    d["Primera"] = pd.to_datetime(d["Primera"], errors="coerce").dt.strftime("%H:%M")
    d["Ultima"] = pd.to_datetime(d["Ultima"], errors="coerce").dt.strftime("%H:%M")

    cols = [
        "Fecha", "Tipo_dia", "Primera", "Ultima",
        "Horas", "Esperado", "Saldo",
        "Marcaciones", "Pares_estimados", "Cortes", "Incompleto", "Cumple"
    ]
    cols = [c for c in cols if c in d.columns]
    return d[cols].sort_values("Fecha").reset_index(drop=True)


def pretty_summary(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    out = df.copy()
    if "EmployeeKey" in out.columns:
        out = out.drop(columns=["EmployeeKey"])
    if "DNI" in out.columns:
        out["DNI"] = out["DNI"].apply(display_dni)
    return out


# =========================
# EXPORT EXCEL
# =========================
def _safe_table_name(name: str) -> str:
    base = re.sub(r"[^A-Za-z0-9_]", "_", name)
    if not base:
        base = "Tabla"
    if not re.match(r"^[A-Za-z_]", base):
        base = f"_{base}"
    return base[:255]


def _apply_excel_style(ws, table_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="0A58CA")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center")

    ws.freeze_panes = "A2"

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row >= 2 and max_col >= 1:
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        ws.auto_filter.ref = ref
        tab = Table(displayName=_safe_table_name(table_name), ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 0
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=col).value
            s = "" if v is None else str(v)
            max_len = max(max_len, len(s))
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)


def export_general_excel(
    reduced: bool,
    driver_mode: bool,
    holidays_text: str,
    expected: int,
    kpis_general: dict,
    summary_all: pd.DataFrame,
    extras_only: pd.DataFrame,
    daily: pd.DataFrame,
    raw: pd.DataFrame,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    def add_df(sheet_name: str, df: pd.DataFrame, text_cols: set[str] | None = None):
        text_cols = text_cols or set()
        ws = wb.create_sheet(sheet_name)

        for j, col in enumerate(df.columns, start=1):
            ws.cell(row=1, column=j, value=str(col))

        for i, row in enumerate(df.itertuples(index=False), start=2):
            for j, val in enumerate(row, start=1):
                col_name = df.columns[j - 1]
                cell = ws.cell(row=i, column=j, value=val)
                if col_name in text_cols and val is not None:
                    cell.value = str(val)
                    cell.number_format = "@"

        _apply_excel_style(ws, table_name=sheet_name)

    df_kpis = pd.DataFrame([{
        "Horario_reducido": "SI" if reduced else "NO",
        "Control_choferes": "SI" if driver_mode else "NO",
        "Esperado_NO_Docente": minutes_to_hhmm(expected),
        "Feriados_cargados": holidays_text.strip(),
        **kpis_general
    }])
    add_df("KPIs_General", df_kpis)

    summary_out = summary_all.copy()
    if "DNI" in summary_out.columns:
        summary_out["DNI"] = summary_out["DNI"].astype(str)
    add_df("Resumen_Empleados", summary_out, text_cols={"DNI", "EmployeeKey"})

    extras_out = extras_only.copy()
    if not extras_out.empty and "DNI" in extras_out.columns:
        extras_out["DNI"] = extras_out["DNI"].astype(str)
    add_df(
        "Solo_Extras",
        extras_out if not extras_out.empty else pd.DataFrame(columns=["Empleado", "DNI", "Tipo", "Horas_extras", "Extras_min"]),
        text_cols={"DNI"}
    )

    daily_out = daily.copy()
    daily_out["DNI"] = daily_out["DNI"].astype(str)
    daily_out["Fecha"] = pd.to_datetime(daily_out["Fecha"]).dt.strftime("%Y-%m-%d")
    daily_out["Primera"] = pd.to_datetime(daily_out["Primera"], errors="coerce").dt.strftime("%Y-%m-%d %H:%M")
    daily_out["Ultima"] = pd.to_datetime(daily_out["Ultima"], errors="coerce").dt.strftime("%Y-%m-%d %H:%M")
    add_df("Detalle_Diario", daily_out, text_cols={"DNI", "EmployeeKey"})

    raw_out = raw.copy()
    raw_out["DNI"] = raw_out["DNI"].astype(str)
    raw_out["Fecha"] = pd.to_datetime(raw_out["FechaHora"]).dt.strftime("%Y-%m-%d")
    raw_out["Hora"] = pd.to_datetime(raw_out["FechaHora"]).dt.strftime("%H:%M")
    raw_out = raw_out.sort_values(["Empleado", "DNI", "FechaHora"])[["EmployeeKey", "Empleado", "DNI", "Tipo", "Fecha", "Hora"]]
    add_df("Marcaciones", raw_out, text_cols={"DNI", "EmployeeKey"})

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# =========================
# ESTADÍSTICAS
# =========================
def safe_pct(a: int, b: int) -> str:
    if b <= 0:
        return "0%"
    return f"{(a / b * 100):.0f}%"


def histogram_hours(series_minutes: pd.Series, bin_hours: list[tuple[float, float]]) -> pd.DataFrame:
    hours = series_minutes.fillna(0).astype(int) / 60.0
    rows = []
    for lo, hi in bin_hours:
        label = f"{lo:.0f}-{hi:.0f}H" if hi < 999 else f"{lo:.0f}H+"
        if hi >= 999:
            cnt = int((hours >= lo).sum())
        else:
            cnt = int(((hours >= lo) & (hours < hi)).sum())
        rows.append({"Rango": label, "Días": cnt})
    return pd.DataFrame(rows)


# =========================
# APP
# =========================
def main() -> None:
    st.set_page_config(page_title="CONTROL DE ASISTENCIA APP", page_icon="🫧", layout="wide")
    inject_css()
    hero_header()
    init_holidays_state()

    with st.container():
        st.markdown('<div class="toolbar-wrap">', unsafe_allow_html=True)
        col1, col2, col3 = st.columns([1, 1, 1])
        with col1:
            reduced = st.toggle("HORARIO REDUCIDO", value=False)
        with col2:
            driver_mode = st.toggle("CONTROL DE CHOFERES", value=False)
        with col3:
            st.markdown(
                f"""<div class="pill">JORNADA {"06:00" if reduced else "07:00"}</div>""",
                unsafe_allow_html=True,
            )
        st.markdown('</div>', unsafe_allow_html=True)

    with st.expander("FERIADOS", expanded=False):
        st.text_area(
            "CARGAR FERIADOS (UNO POR LÍNEA O SEPARADOS POR COMA · FORMATO DD/MM/AAAA O AAAA-MM-DD)",
            key="holidays_text",
            height=110,
        )

        a1, a2, a3 = st.columns(3)
        with a1:
            if st.button("APLICAR FERIADOS ESCRITOS", use_container_width=True):
                apply_text_holidays()
                st.rerun()
        with a2:
            if st.button("LIMPIAR TODOS LOS FERIADOS", use_container_width=True):
                clear_all_holidays()
                st.rerun()
        with a3:
            st.markdown(
                f"""<div class="pill">FERIADOS CARGADOS: {len(st.session_state["holidays_set"])}</div>""",
                unsafe_allow_html=True,
            )

        render_holiday_calendar()

        if st.session_state["holidays_set"]:
            holidays_df = pd.DataFrame(
                {"FERIADO": [d.strftime("%d/%m/%Y") for d in sorted(st.session_state["holidays_set"])]}
            )
            st.dataframe(holidays_df, use_container_width=True, height=180, hide_index=True)

    holidays = set(st.session_state["holidays_set"])
    holidays_text = st.session_state["holidays_text"]

    file = st.file_uploader("", type=["xlsx", "xlsm", "xls"], label_visibility="collapsed")
    if not file:
        return

    df0 = read_excel_auto(file)
    df0 = validate_format(df0)
    raw0 = parse_and_clean(df0)
    _ = init_profiles(raw0)

    expected = 360 if reduced else 420
    tabs = st.tabs(["GENERAL", "EMPLEADO", "PERFILES"])

    # =======================
    # GENERAL
    # =======================
    with tabs[0]:
        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        raw = apply_profiles(raw0, st.session_state["profiles"])

        fixes_total = 0
        if not driver_mode:
            left, right = st.columns([1.05, 1])
            with left:
                fix_all = st.button("CORREGIR FALTAS DE MARCACIÓN (TODOS)", use_container_width=True)
            with right:
                st.markdown("""<div class="pill">SOLO NO DOCENTES CON 1 MARCACIÓN</div>""", unsafe_allow_html=True)

            if fix_all:
                raw, fixes_total = correct_missing_punches_all(raw, expected)
                st.markdown(f"""<div class="pill">CORRECCIONES APLICADAS: {fixes_total}</div>""", unsafe_allow_html=True)

        daily = calc_daily(raw, expected, holidays, driver_mode)
        summary = summarize(daily)

        total_min = int(daily["Minutos"].sum()) if not daily.empty else 0
        empleados = int(summary.shape[0]) if not summary.empty else 0
        dias = int(daily["Fecha"].nunique()) if not daily.empty else 0
        prom_dia = int(round(daily["Minutos"].mean())) if not daily.empty else 0

        total_marc = int(daily["Marcaciones"].sum()) if not daily.empty else 0
        incompletos = int((daily["Incompleto"] == "SI").sum()) if not daily.empty else 0
        cortes = int((daily["Cortes"] == "SI").sum()) if not daily.empty else 0
        total_registros_dia = int(daily.shape[0]) if not daily.empty else 0

        nod = daily[daily["Tipo"] == "NO Docente"].copy()
        nod_sum = int(nod["Minutos"].sum()) if not nod.empty else 0
        nod_exp_sum = int(nod["Esperado_min"].sum()) if not nod.empty else 0
        nod_pct = f"{(nod_sum / nod_exp_sum * 100):.0f}%" if nod_exp_sum > 0 else ""

        nod_extras = int(nod.loc[nod["Saldo_min"] > 0, "Saldo_min"].sum()) if not nod.empty else 0
        nod_faltas = int((-nod.loc[nod["Saldo_min"] < 0, "Saldo_min"].sum())) if not nod.empty else 0

        doc = daily[daily["Tipo"] == "Docente"].copy()
        doc_sum = int(doc["Minutos"].sum()) if not doc.empty else 0

        r1 = st.columns(4)
        with r1[0]:
            kpi_card("EMPLEADOS", f"{empleados}", f"DÍAS: {dias}")
        with r1[1]:
            kpi_card("TOTAL", minutes_to_hhmm(total_min), f"PROM/DÍA: {minutes_to_hhmm(prom_dia)}")
        with r1[2]:
            kpi_card("MARCACIONES", f"{total_marc}", f"INCOMPLETOS: {incompletos}")
        with r1[3]:
            kpi_card("CORTES", f"{cortes}", safe_pct(cortes, total_registros_dia))

        r2 = st.columns(4)
        with r2[0]:
            kpi_card("NO DOCENTE", minutes_to_hhmm(nod_sum), f"CUMPLIMIENTO: {nod_pct}")
        with r2[1]:
            kpi_card("EXTRAS NO DOCENTE", minutes_to_hhmm(nod_extras), "INCLUYE FERIADOS Y FIN DE SEMANA")
        with r2[2]:
            kpi_card("FALTAS NO DOCENTE", minutes_to_hhmm(nod_faltas), "SOLO DÍAS HÁBILES")
        with r2[3]:
            kpi_card("DOCENTE", minutes_to_hhmm(doc_sum), "POR TRAMOS")

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        extras_only = summary[summary["Tipo"] == "NO Docente"].copy()
        if not extras_only.empty:
            extras_only = (
                extras_only[["Empleado", "DNI", "Tipo", "Extras", "Extras_min"]]
                .rename(columns={"Extras": "Horas_extras"})
                .sort_values("Extras_min", ascending=False)
                .reset_index(drop=True)
            )
            extras_only["DNI"] = extras_only["DNI"].apply(display_dni)
        else:
            extras_only = pd.DataFrame(columns=["Empleado", "DNI", "Tipo", "Horas_extras", "Extras_min"])

        copy_table_button(extras_only, "COPIAR SOLO EXTRAS", key="copy_extras")
        st.table(extras_only)

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        kpis_general = {
            "Empleados": empleados,
            "Dias_total": dias,
            "Total_HHMM": minutes_to_hhmm(total_min),
            "Prom_dia_HHMM": minutes_to_hhmm(prom_dia),
            "Incompletos": incompletos,
            "Cortes": cortes,
            "NO_Docente_Total_HHMM": minutes_to_hhmm(nod_sum),
            "NO_Docente_Cumplimiento": nod_pct,
            "NO_Docente_Extras_HHMM": minutes_to_hhmm(nod_extras),
            "NO_Docente_Faltas_HHMM": minutes_to_hhmm(nod_faltas),
            "Correcciones_masivas": fixes_total,
        }

        general_xlsx = export_general_excel(
            reduced=reduced,
            driver_mode=driver_mode,
            holidays_text=holidays_text,
            expected=expected,
            kpis_general=kpis_general,
            summary_all=summary.copy(),
            extras_only=extras_only.copy(),
            daily=daily.copy(),
            raw=raw.copy(),
        )

        st.download_button(
            "EXPORTAR RESUMEN GENERAL (EXCEL)",
            data=general_xlsx,
            file_name="resumen_general_asistencia.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        if not daily.empty:
            by_day = daily.groupby("Fecha", as_index=False).agg(Minutos=("Minutos", "sum"))
            by_day["Horas"] = by_day["Minutos"] / 60.0
            by_day = by_day.sort_values("Fecha")
            st.line_chart(by_day.set_index("Fecha")[["Horas"]], height=230)

            hist = histogram_hours(daily["Minutos"], [(0, 2), (2, 4), (4, 6), (6, 8), (8, 10), (10, 999)]).set_index("Rango")
            st.bar_chart(hist[["Días"]], height=230)

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        summary_show = pretty_summary(summary)
        copy_table_button(summary_show, "COPIAR RESUMEN COMPLETO", key="copy_summary")
        st.dataframe(summary_show, use_container_width=True, height=560, hide_index=True)

        st.session_state["__raw__"] = raw
        st.session_state["__daily__"] = daily
        st.session_state["__summary__"] = summary
        st.session_state["__expected__"] = expected
        st.session_state["__driver_mode__"] = driver_mode
        st.session_state["__holidays__"] = holidays

    # =======================
    # EMPLEADO
    # =======================
    with tabs[1]:
        raw = st.session_state.get("__raw__", None)
        daily = st.session_state.get("__daily__", None)
        summary = st.session_state.get("__summary__", None)
        expected = st.session_state.get("__expected__", expected)
        driver_mode = st.session_state.get("__driver_mode__", driver_mode)
        holidays = st.session_state.get("__holidays__", holidays)

        if raw is None or daily is None or summary is None or summary.empty:
            st.info("CARGÁ UN EXCEL PARA VER ESTA SECCIÓN.")
            return

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        disp = summary.copy()
        disp["Display"] = (
            disp["Empleado"].astype(str)
            + " · "
            + disp["DNI"].apply(display_dni)
            + " · "
            + disp["Tipo"].astype(str)
        )
        selected = st.selectbox("", options=disp["Display"].tolist(), label_visibility="collapsed")
        r = disp[disp["Display"] == selected].iloc[0]

        ekey = r["EmployeeKey"]
        emp = r["Empleado"]
        dni = str(r["DNI"])
        tipo = r["Tipo"]

        raw_emp = raw[(raw["EmployeeKey"] == ekey) & (raw["Tipo"] == tipo)].copy().sort_values("FechaHora")

        fixes_applied = 0
        if tipo == "NO Docente" and not driver_mode:
            fix = st.button("CORREGIR FALTA DE MARCACIÓN", use_container_width=True)
            if fix:
                corrected_raw_emp, fixes_applied = correct_missing_punches_for_employee(raw_emp, expected)

                raw_corrected = raw.copy()
                mask = (raw_corrected["EmployeeKey"] == ekey) & (raw_corrected["Tipo"] == tipo)
                raw_corrected = raw_corrected[~mask]
                raw_corrected = pd.concat([raw_corrected, corrected_raw_emp], ignore_index=True)
                raw_corrected = raw_corrected.sort_values(["Empleado", "DNI", "FechaHora"]).reset_index(drop=True)

                daily_corrected = calc_daily(raw_corrected, expected, holidays, driver_mode)
                daily_emp = daily_corrected[
                    (daily_corrected["EmployeeKey"] == ekey) &
                    (daily_corrected["Tipo"] == tipo)
                ].copy()
            else:
                daily_emp = daily[
                    (daily["EmployeeKey"] == ekey) &
                    (daily["Tipo"] == tipo)
                ].copy()
        else:
            daily_emp = daily[
                (daily["EmployeeKey"] == ekey) &
                (daily["Tipo"] == tipo)
            ].copy()

        if daily_emp.empty:
            st.warning("SIN DATOS PARA ESTE EMPLEADO.")
            return

        total_min = int(daily_emp["Minutos"].sum())
        dias = int(daily_emp["Fecha"].nunique())
        prom = int(round(daily_emp["Minutos"].mean())) if dias else 0

        inc = int((daily_emp["Incompleto"] == "SI").sum())
        cuts = int((daily_emp["Cortes"] == "SI").sum())
        marc_total = int(daily_emp["Marcaciones"].sum())
        marc_prom = marc_total / max(int(daily_emp.shape[0]), 1)

        exp_sum = int(daily_emp["Esperado_min"].sum())
        saldo_sum = int(daily_emp["Saldo_min"].sum())

        extras_min = 0
        faltas_min = 0
        pct = ""
        if tipo == "NO Docente":
            extras_min = int(daily_emp.loc[daily_emp["Saldo_min"] > 0, "Saldo_min"].sum())
            faltas_min = int((-daily_emp.loc[daily_emp["Saldo_min"] < 0, "Saldo_min"].sum()))
            pct = f"{(total_min / exp_sum * 100):.0f}%" if exp_sum > 0 else ""

        r1 = st.columns(4)
        with r1[0]:
            kpi_card(emp, minutes_to_hhmm(total_min), f"{tipo} · {display_dni(dni)}")
        with r1[1]:
            kpi_card("DÍAS", f"{dias}", f"PROM/DÍA: {minutes_to_hhmm(prom)}")
        with r1[2]:
            kpi_card("MARCACIONES", f"{marc_total}", f"PROM/DÍA: {marc_prom:.2f}")
        with r1[3]:
            if tipo == "NO Docente":
                kpi_card("EXTRAS", minutes_to_hhmm(extras_min), f"FALTAS: {minutes_to_hhmm(faltas_min)} · {pct}")
            else:
                kpi_card("TOTAL", minutes_to_hhmm(total_min), "DOCENTE")

        r2 = st.columns(3)
        with r2[0]:
            kpi_card("INCOMPLETOS", f"{inc}", "")
        with r2[1]:
            kpi_card("CORTES", f"{cuts}", "")
        with r2[2]:
            if tipo == "NO Docente":
                kpi_card("SALDO", delta_short(saldo_sum), "ACUMULADO")
            else:
                kpi_card("SALDO", "—", "")

        if fixes_applied:
            st.markdown(f"""<div class="pill">CORRECCIONES APLICADAS: {fixes_applied}</div>""", unsafe_allow_html=True)

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        ch = daily_emp.sort_values("Fecha").copy()
        ch["Horas_float"] = ch["Minutos"] / 60.0
        st.bar_chart(ch.set_index("Fecha")[["Horas_float"]], height=240)

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        det = employee_detail_table(daily_emp)
        copy_table_button(det, "COPIAR DETALLE DÍA A DÍA", key="copy_emp_detail")
        st.dataframe(det, use_container_width=True, height=560, hide_index=True)

    # =======================
    # PERFILES
    # =======================
    with tabs[2]:
        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        profiles_view = st.session_state["profiles"].copy()
        profiles_view["DNI"] = profiles_view["DNI"].apply(display_dni)

        edited = st.data_editor(
            profiles_view,
            use_container_width=True,
            height=600,
            hide_index=True,
            disabled=["EmployeeKey", "DNI", "Empleado"],
            column_config={
                "EmployeeKey": st.column_config.TextColumn("ID INTERNO", width="medium"),
                "DNI": st.column_config.TextColumn("DNI", width="small"),
                "Empleado": st.column_config.TextColumn("EMPLEADO", width="large"),
                "Tipo": st.column_config.SelectboxColumn("TIPO", options=["NO Docente", "Docente"], required=True),
            },
        )

        real_profiles = st.session_state["profiles"].copy()
        real_profiles["Tipo"] = edited["Tipo"]
        st.session_state["profiles"] = real_profiles


if __name__ == "__main__":
    main()
    
